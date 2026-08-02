# spider_zhuwang/pig_spider.py
"""
生猪价格爬虫 - 完整版
包含下载、合并、异常检测等完整功能
"""

import requests
import pandas as pd
import numpy as np
import os
import time
import re
from typing import Optional, List, Tuple, Dict
from datetime import datetime
from bs4 import BeautifulSoup
import warnings
import scipy.stats as stats
warnings.filterwarnings('ignore')

# 获取当前文件所在目录
_CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
# 项目根目录
_PROJECT_ROOT = os.path.dirname(_CURRENT_DIR)
# 默认保存目录
DEFAULT_SAVE_DIR = os.path.join(_PROJECT_ROOT, 'data_save', 'yangzhuwang_shengzhu')

# ==================== 配置 ====================
class PigSpiderConfig:
    """生猪爬虫配置"""
    DEFAULT_SAVE_DIR = DEFAULT_SAVE_DIR
    DEFAULT_CREATOR = "mzj"
    USER_AGENT = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    TIMEOUT = 30
    RETRY_TIMES = 3
    RETRY_DELAY = 2


class PigPriceSpider:
    """
    生猪价格数据爬虫
    支持从中国养猪网下载价格数据，转换为宽格式（日期为行，省份为列）
    """
    
    def __init__(self, save_dir: Optional[str] = None, creator: Optional[str] = None):
        self.save_dir = save_dir or PigSpiderConfig.DEFAULT_SAVE_DIR
        self.creator = creator or PigSpiderConfig.DEFAULT_CREATOR
        self.headers = {"User-Agent": PigSpiderConfig.USER_AGENT}
        os.makedirs(self.save_dir, exist_ok=True)
        print(f"🐖 生猪价格爬虫初始化完成")
        print(f"📁 保存目录: {self.save_dir}")
    
    # ==================== 获取最新合并文件 ====================
    def get_latest_merge_file(self) -> Optional[str]:
        import glob
        pattern = os.path.join(self.save_dir, "合并标记_*.xlsx")
        files = glob.glob(pattern)
        if not files:
            return None
        return max(files, key=os.path.getmtime)
    
    # ==================== 从列表页获取链接 ====================
    def get_links_from_list_page(self, list_url: str, max_pages: Optional[int] = None, 
                                  retry_times: int = 3, retry_delay: int = 2) -> List[str]:
        print(f"开始从列表页抓取链接: {list_url}")
        all_urls = []
        failed_pages = []
        current_page = 1
        
        match = re.match(r'(.*list-\d+-)\d+(\.html)', list_url)
        if not match:
            print(f"错误: 无法解析列表页URL格式: {list_url}")
            return []
        
        base_url_prefix = match.group(1)
        base_url_suffix = match.group(2)
        
        while True:
            page_url = f"{base_url_prefix}{current_page}{base_url_suffix}"
            if current_page == 1:
                page_url = list_url
            
            print(f"  正在抓取第 {current_page} 页: {page_url}")
            
            success = False
            page_urls = []
            
            for attempt in range(1, retry_times + 1):
                try:
                    response = requests.get(page_url, headers=self.headers, timeout=PigSpiderConfig.TIMEOUT)
                    response.raise_for_status()
                    response.encoding = 'utf-8'
                    soup = BeautifulSoup(response.text, 'html.parser')
                    
                    zxleft3 = soup.find('div', class_='zxleft3')
                    if not zxleft3:
                        success = True
                        break
                    
                    ul = zxleft3.find('ul')
                    if not ul:
                        success = True
                        break
                    
                    for li in ul.find_all('li'):
                        a_tag = li.find('a')
                        if not a_tag:
                            continue
                        href = a_tag.get('href')
                        if not href or '/shengzhu/' not in href:
                            continue
                        
                        if href.startswith('/'):
                            full_url = f"https://hangqing.zhuwang.com.cn{href}"
                        elif href.startswith('http'):
                            full_url = href
                        else:
                            full_url = re.sub(r'/[^/]*$', '/', list_url) + href
                        
                        page_urls.append(full_url)
                    
                    success = True
                    break
                    
                except Exception as e:
                    print(f"  第 {attempt} 次请求失败: {e}")
                    if attempt < retry_times:
                        time.sleep(retry_delay)
            
            if success:
                all_urls.extend(page_urls)
                print(f"  第 {current_page} 页抓取完成，共 {len(page_urls)} 条链接")
            else:
                print(f"  ✗ 第 {current_page} 页抓取失败，已跳过")
                failed_pages.append(current_page)
            
            if max_pages is not None and current_page >= max_pages:
                break
            
            if success:
                try:
                    response = requests.get(page_url, headers=self.headers, timeout=10)
                    response.encoding = 'utf-8'
                    soup = BeautifulSoup(response.text, 'html.parser')
                    pagination = soup.find('div', class_='zxpage')
                    if pagination:
                        next_link = pagination.find('a', string='下一页')
                        if next_link and next_link.get('href'):
                            current_page += 1
                            time.sleep(1)
                            continue
                except:
                    pass
            
            if not success:
                current_page += 1
                if len(failed_pages) > 5:
                    print(f"  连续失败次数过多，停止抓取")
                    break
                continue
            
            break
        
        print(f"\n✓ 共抓取 {len(all_urls)} 个链接")
        if failed_pages:
            print(f"  ⚠ 失败页码: {failed_pages}")
        return all_urls
    
    # ==================== 下载单日数据 ====================
    def download_single(self, detail_url: str, merge_path: Optional[str] = None) -> Optional[str]:
        if merge_path is None:
            merge_path = self.get_latest_merge_file()
            if merge_path:
                print(f"  自动选择最新合并文件: {os.path.basename(merge_path)}")
        
        print(f"开始处理: {detail_url}")
        
        date_str, code = self._extract_date_and_code(detail_url)
        if not date_str or not code:
            print(f"错误: 无法从链接提取日期或编码")
            return None
        
        soup = self._get_soup(detail_url)
        if not soup:
            return None
        
        variety = self._extract_variety(soup)
        if not variety:
            print(f"错误: 无法从页面提取品种")
            return None
        print(f"  识别品种: {variety}")
        
        df_wide = self._parse_table(soup, date_str)
        if df_wide is None or df_wide.empty:
            print("错误: 未能提取到有效数据")
            return None
        
        df_wide = self._add_metadata(df_wide, variety, code)
        df_wide = df_wide.reset_index().rename(columns={'index': '日期'})
        
        if merge_path and os.path.exists(merge_path):
            print(f"\n🔄 正在合并到主数据文件...")
            result_path = self._merge_to_main(df_wide, variety, date_str, code, merge_path)
            if result_path:
                print(f"✅ 合并完成: {result_path}")
                return result_path
        
        filename = f"{variety}_{date_str}_{code}.xlsx"
        filepath = os.path.join(self.save_dir, filename)
        df_wide.to_excel(filepath, sheet_name='数据', index=False)
        print(f"✓ 数据已保存: {filepath}")
        return filepath
    
    # ==================== ✅ 核心：从列表页下载 ====================
    # ==================== ✅ 在这里添加新方法 ====================
    def download_from_list_page(self, list_url: str, max_pages: Optional[int] = None, 
                                merge_path: Optional[str] = None) -> Optional[pd.DataFrame]:
        """
        从列表页自动抓取所有链接并批量下载，自动合并到主数据文件
        
        参数:
            list_url: 列表页URL
            max_pages: 最大抓取页数，None表示抓取所有页
            merge_path: 主数据文件路径，默认自动查找最新合并文件
        
        返回:
            合并后的DataFrame
        """
        print(f"📋 从列表页获取链接: {list_url}")
        
        # 1. 从列表页获取所有链接
        urls = self.get_links_from_list_page(list_url, max_pages)
        
        if not urls:
            print("❌ 未获取到任何链接，退出")
            return None
        
        print(f"✅ 获取到 {len(urls)} 个链接")
        print(f"📦 准备批量下载...")
        
        # 2. 批量下载
        result = self.download_batch(urls, merge_path)
        
        return result
    # ==================== 批量下载 ====================
    def download_batch(self, urls: List[str], merge_path: Optional[str] = None) -> Optional[pd.DataFrame]:
        if merge_path is None:
            merge_path = self.get_latest_merge_file()
            if merge_path:
                print(f"  自动选择最新合并文件: {os.path.basename(merge_path)}")
        
        print(f"📦 批量下载，共 {len(urls)} 个链接")
        
        all_data = []
        success_count = 0
        failed_urls = []
        
        for idx, url in enumerate(urls, 1):
            print(f"  [{idx}/{len(urls)}] 下载: {url}")
            result = self._download_to_df(url)
            if result is None:
                failed_urls.append(url)
                continue
            
            df_wide, variety, date_str, code = result
            all_data.append(df_wide)
            success_count += 1
            print(f"    成功下载: {date_str} ({variety})")
            time.sleep(0.5)
        
        if not all_data:
            print("❌ 没有成功下载任何数据")
            return None
        
        df_merged = pd.concat(all_data, ignore_index=True)
        print(f"✅ 成功下载 {success_count}/{len(urls)} 个数据")
        
        if merge_path and os.path.exists(merge_path):
            print(f"\n🔄 正在合并到主数据文件...")
            temp_file = os.path.join(self.save_dir, f"_temp_batch_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
            df_merged.to_excel(temp_file, sheet_name='数据', index=False)
            
            try:
                result_path = self.merge_and_mark_historical(
                    file_a=merge_path,
                    file_b=temp_file
                )
                os.remove(temp_file) if os.path.exists(temp_file) else None
                
                if result_path:
                    print(f"✅ 合并完成: {result_path}")
                    return pd.read_excel(result_path, sheet_name='保留_最新数据')
            except Exception as e:
                print(f"合并失败: {e}")
                os.remove(temp_file) if os.path.exists(temp_file) else None
                return df_merged
        
        return df_merged
    
    # ==================== 辅助方法 ====================
    def _get_soup(self, url: str) -> Optional[BeautifulSoup]:
        try:
            response = requests.get(url, headers=self.headers, timeout=PigSpiderConfig.TIMEOUT)
            response.raise_for_status()
            response.encoding = 'utf-8'
            return BeautifulSoup(response.text, 'html.parser')
        except Exception as e:
            print(f"获取页面内容失败: {e}")
            return None
    
    def _extract_date_and_code(self, url: str) -> Tuple[Optional[str], Optional[str]]:
        pattern = r'/(\d{8})/(\d+)\.html'
        match = re.search(pattern, url)
        if match:
            date_str = match.group(1)
            code = match.group(2)
            return f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:8]}", code
        return None, None
    
    def _extract_variety(self, soup) -> Optional[str]:
        h3_tag = soup.find('h3')
        if h3_tag:
            h3_text = h3_tag.get_text()
            match = re.search(r'全国(外三元|内三元|土杂猪)生猪价格', h3_text)
            if match:
                return match.group(1)
        return None
    
    def _parse_table(self, soup, date_str: str) -> Optional[pd.DataFrame]:
        table = soup.find('table', class_='tabzj')
        if not table:
            return None
        
        tbody = table.find('tbody') or table
        rows_data = []
        current_region = None
        region_rowspan_counter = 0
        
        for tr in tbody.find_all('tr'):
            if '当前内容来源' in tr.get_text():
                continue
            
            tds = tr.find_all('td')
            if not tds:
                continue
            
            row_texts = [td.get_text(separator=' ', strip=True) for td in tds]
            
            first_td = tds[0]
            if first_td.has_attr('rowspan'):
                current_region = row_texts[0]
                try:
                    region_rowspan_counter = int(first_td['rowspan']) - 1
                except:
                    region_rowspan_counter = 0
                row_data = [current_region] + row_texts[1:]
            else:
                if current_region and region_rowspan_counter > 0:
                    row_data = [current_region] + row_texts
                    region_rowspan_counter -= 1
                else:
                    row_data = [""] + row_texts
            
            if len(row_data) >= 3:
                rows_data.append({
                    'region': row_data[0],
                    'province': row_data[1],
                    'price': self._safe_float(row_data[2]),
                })
        
        if not rows_data:
            return None
        
        df_long = pd.DataFrame(rows_data)
        price_dict = df_long.set_index('province')['price'].to_dict()
        return pd.DataFrame([price_dict], index=[date_str])
    
    def _add_metadata(self, df: pd.DataFrame, variety: str, code: str) -> pd.DataFrame:
        df_copy = df.copy()
        df_copy.insert(0, '品种', variety)
        df_copy.insert(1, '编码', code)
        df_copy['创建时间'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        df_copy['创建人'] = self.creator
        return df_copy
    
    def _safe_float(self, value) -> Optional[float]:
        try:
            return float(value) if value else None
        except (ValueError, TypeError):
            return None
    
    def _merge_to_main(self, df_new: pd.DataFrame, variety: str, date_str: str, 
                       code: str, merge_path: str) -> Optional[str]:
        temp_file = os.path.join(self.save_dir, f"_temp_{variety}_{date_str}_{code}.xlsx")
        df_new.to_excel(temp_file, sheet_name='数据', index=False)
        
        try:
            result_path = self.merge_and_mark_historical(
                file_a=merge_path,
                file_b=temp_file
            )
            os.remove(temp_file) if os.path.exists(temp_file) else None
            return result_path
        except Exception as e:
            print(f"合并失败: {e}")
            os.remove(temp_file) if os.path.exists(temp_file) else None
            return None
    
    def _download_to_df(self, url: str) -> Optional[Tuple[pd.DataFrame, str, str, str]]:
        date_str, code = self._extract_date_and_code(url)
        if not date_str or not code:
            return None
        
        soup = self._get_soup(url)
        if not soup:
            return None
        
        variety = self._extract_variety(soup)
        if not variety:
            return None
        
        df_wide = self._parse_table(soup, date_str)
        if df_wide is None or df_wide.empty:
            return None
        
        df_wide = self._add_metadata(df_wide, variety, code)
        df_wide = df_wide.reset_index().rename(columns={'index': '日期'})
        
        return df_wide, variety, date_str, code
    
    # ==================== 数据合并和历史标记 ====================
    def merge_and_mark_historical(self, file_a: str, file_b: str,
                                   output_filename: Optional[str] = None,
                                   date_col: str = '日期',
                                   variety_col: str = '品种',
                                   code_col: str = '编码',
                                   create_time_col: str = '创建时间') -> Optional[str]:
        print(f"=== 数据合并与历史标记 ===")
        print(f"原始数据文件(A): {file_a}")
        print(f"新增数据文件(B): {file_b}")
        
        df_a = self._read_data_file(file_a, [date_col, variety_col, code_col, create_time_col])
        df_b = self._read_data_file(file_b, [date_col, variety_col, code_col, create_time_col])
        
        if df_a is None or df_a.empty:
            print("错误: A文件为空或读取失败")
            return None
        if df_b is None or df_b.empty:
            print("错误: B文件为空或读取失败")
            return None
        
        print(f"  A文件: {len(df_a)} 行")
        print(f"  B文件: {len(df_b)} 行")
        
        # 合并
        df_merged = pd.concat([df_a, df_b], ignore_index=True)
        print(f"  合并后总行数: {len(df_merged)}")
        
        # 预处理
        df_copy = df_merged.copy()
        
        if date_col in df_copy.columns:
            df_copy[date_col] = df_copy[date_col].apply(
                lambda x: x.strftime('%Y-%m-%d') if isinstance(x, pd.Timestamp) else str(x).strip()
            )
        
        if create_time_col in df_copy.columns:
            df_copy['_create_time_parsed'] = pd.to_datetime(df_copy[create_time_col], errors='coerce')
        
        key_cols = [date_col, variety_col, code_col]
        df_copy['_group_key'] = df_copy[key_cols].apply(
            lambda row: '|||'.join(row.astype(str).str.strip()), axis=1
        )
        
        df_copy['是否保留'] = '保留'
        group_counts = df_copy['_group_key'].value_counts()
        duplicate_keys = group_counts[group_counts > 1].index.tolist()
        
        history_count = 0
        for group_key in duplicate_keys:
            group_indices = df_copy[df_copy['_group_key'] == group_key].index.tolist()
            if len(group_indices) > 1:
                group_data = df_copy.loc[group_indices]
                sorted_indices = group_data.sort_values('_create_time_parsed', ascending=False).index.tolist()
                history_indices = sorted_indices[1:]
                df_copy.loc[history_indices, '是否保留'] = '历史'
                history_count += len(history_indices)
        
        keep_count = len(df_copy[df_copy['是否保留'] == '保留'])
        print(f"  保留: {keep_count} 条, 历史: {history_count} 条")
        
        df_result = df_copy.drop(columns=['_group_key', '_create_time_parsed'], errors='ignore')
        df_result['数据来源'] = ''
        df_result.loc[:len(df_a)-1, '数据来源'] = '原始数据(A)'
        df_result.loc[len(df_a):, '数据来源'] = '新增数据(B)'
        
        df_result[date_col] = pd.to_datetime(df_result[date_col], errors='coerce').dt.strftime('%Y-%m-%d')
        df_result = df_result.sort_values(by=[date_col, code_col], ascending=[False, False]).reset_index(drop=True)
        
        if output_filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_filename = f"合并标记_{timestamp}.xlsx"
        
        output_path = os.path.join(self.save_dir, output_filename)
        
        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df_result.to_excel(writer, sheet_name='合并标记结果', index=False)
                
                keep_df = df_result[df_result['是否保留'] == '保留'].copy()
                if not keep_df.empty:
                    exclude_cols = ['日期', '品种', '编码', '创建时间', '创建人', '是否保留', '数据来源']
                    numeric_cols = [col for col in keep_df.columns if col not in exclude_cols]
                    for col in numeric_cols:
                        keep_df[col] = pd.to_numeric(keep_df[col], errors='coerce')
                    
                    keep_df['平均值'] = keep_df[numeric_cols].mean(axis=1, skipna=True)
                    keep_df['标准差'] = keep_df[numeric_cols].std(axis=1, skipna=True, ddof=0)
                    keep_df['最大值'] = keep_df[numeric_cols].max(axis=1, skipna=True)
                    keep_df['最小值'] = keep_df[numeric_cols].min(axis=1, skipna=True)
                    
                    other_cols = [col for col in keep_df.columns if col not in ['日期', '品种', '编码', '平均值', '标准差', '最大值', '最小值', '创建时间', '创建人', '是否保留', '数据来源']]
                    final_cols = ['日期', '品种', '编码', '平均值', '标准差', '最大值', '最小值'] + other_cols + ['创建时间', '创建人', '是否保留', '数据来源']
                    keep_df = keep_df[final_cols]
                    keep_df.to_excel(writer, sheet_name='保留_最新数据', index=False)
                
                history_df = df_result[df_result['是否保留'] == '历史'].copy()
                history_df.to_excel(writer, sheet_name='历史_需归档', index=False)
                
                stats_df = pd.DataFrame({
                    '统计项': ['A文件行数', 'B文件行数', '合并总行数', '保留记录数', '历史记录数'],
                    '值': [len(df_a), len(df_b), len(df_result), keep_count, history_count]
                })
                stats_df.to_excel(writer, sheet_name='统计信息', index=False)
            
            print(f"✓ 合并标记结果已保存: {output_path}")
            return output_path
            
        except Exception as e:
            print(f"保存文件失败: {e}")
            return None
    
    def _read_data_file(self, file_path: str, required_cols: List[str]) -> Optional[pd.DataFrame]:
        try:
            xls = pd.ExcelFile(file_path)
            for sheet in xls.sheet_names:
                if sheet in ['统计信息', '下载记录', '失败记录', '保留_最新数据', '历史_需归档']:
                    continue
                df = pd.read_excel(file_path, sheet_name=sheet)
                if len(df) > 0 and any(col in df.columns for col in required_cols):
                    return df
            return pd.read_excel(file_path, sheet_name=0)
        except Exception as e:
            print(f"读取文件失败: {e}")
            return None
    
    # ==================== 3σ异常值检测 ====================
    def mark_anomalies_by_3sigma(self, file_path: str,
                                  date_col: str = '日期',
                                  variety_col: str = '品种',
                                  code_col: str = '编码',
                                  mean_col: str = '平均值',
                                  std_col: str = '标准差',
                                  sheet_name: str = '保留_最新数据') -> Optional[str]:
        print(f"=== 3σ原则异常值标记 ===")
        print(f"数据文件: {file_path}")
        
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            print(f"  读取数据: {len(df)} 行")
        except Exception as e:
            print(f"读取文件失败: {e}")
            return None
        
        required_cols = [date_col, variety_col, code_col, mean_col, std_col]
        for col in required_cols:
            if col not in df.columns:
                print(f"错误: 缺少必要字段 '{col}'")
                return None
        
        exclude_cols = [date_col, variety_col, code_col, mean_col, std_col, 
                        '创建时间', '创建人', '是否保留', '数据来源', '最大值', '最小值']
        province_cols = [col for col in df.columns if col not in exclude_cols]
        print(f"  识别到 {len(province_cols)} 个省份列")
        
        anomaly_records = []
        
        for idx, row in df.iterrows():
            mean_val = row[mean_col]
            std_val = row[std_col]
            if pd.isna(std_val) or std_val == 0:
                continue
            
            lower_bound = mean_val - 3 * std_val
            upper_bound = mean_val + 3 * std_val
            
            for col in province_cols:
                val = row[col]
                if pd.isna(val):
                    continue
                try:
                    val = float(val)
                except:
                    continue
                
                if val < lower_bound or val > upper_bound:
                    anomaly_records.append({
                        '行号': idx + 2,
                        '日期': row[date_col],
                        '品种': row[variety_col],
                        '编码': row[code_col],
                        '省份': col,
                        '价格': val,
                        '偏差类型': "偏大" if val >= upper_bound else "偏小", # 偏大还是偏小
                        '平均值': mean_val,
                        '标准差': std_val,
                        '下限': lower_bound,
                        '上限': upper_bound,
                        '偏差': f"{(val - mean_val):.2f}",
                        '偏差倍数': f"{(val - mean_val) / std_val:.2f}σ",
                        '置信度': f"{1 - 2 * (1 - stats.norm.cdf(abs((val - mean_val) / std_val))):.10%}",
                        '异常概率': f"{2 * (1 - stats.norm.cdf(abs((val - mean_val) / std_val))):.10%}",
                    })
        
        if not anomaly_records:
            print("  ✓ 未发现异常值")
            return file_path
        
        print(f"  发现 {len(anomaly_records)} 个异常值")
        
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font
            
            workbook = load_workbook(file_path)
            worksheet = workbook[sheet_name]
            col_indices = {col_name: idx + 1 for idx, col_name in enumerate(df.columns)}
            red_font = Font(color="FF0000")
            
            for record in anomaly_records:
                col_idx = col_indices.get(record['省份'])
                if col_idx:
                    cell = worksheet.cell(row=record['行号'], column=col_idx)
                    cell.font = red_font
            
            anomaly_df = pd.DataFrame(anomaly_records)
            if '异常值汇总' in workbook.sheetnames:
                workbook.remove(workbook['异常值汇总'])
            ws = workbook.create_sheet('异常值汇总')
            for r_idx, row in enumerate([anomaly_df.columns.tolist()] + anomaly_df.values.tolist(), 1):
                for c_idx, val in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=val)
            
            workbook.save(file_path)
            print(f"\n✓ 3σ异常值标记完成，已标记 {len(anomaly_records)} 个异常值")
            return file_path
            
        except Exception as e:
            print(f"处理文件失败: {e}")
            return None